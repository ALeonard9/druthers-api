"""Generic CSV/XLSX movie history import with pre-ingestion validation."""

import csv
import io
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session

from app.db.models_sandbox import DbMovie, DbUserMovie
from app.services.movie_search import apply_detail_to_movie, get_movie_detail
from app.services.tracker_rules import utc_now

MOVIE_TEMPLATE_COLUMNS = (
    {
        'name': 'title',
        'format': 'text, maximum 255 characters',
        'description': 'Movie title. Must match the movie identified by tmdb_id.',
        'example': 'The Matrix',
    },
    {
        'name': 'release_year',
        'format': 'four-digit year (YYYY)',
        'description': 'Original release year. Must match the TMDB movie.',
        'example': '1999',
    },
    {
        'name': 'tmdb_id',
        'format': 'positive integer',
        'description': 'TMDB movie ID, found in a themoviedb.org movie URL.',
        'example': '603',
    },
    {
        'name': 'watched_date',
        'format': 'ISO date (YYYY-MM-DD), not in the future',
        'description': 'Date the movie was watched.',
        'example': '2026-08-15',
    },
)
MOVIE_TEMPLATE_HEADERS = tuple(column['name'] for column in MOVIE_TEMPLATE_COLUMNS)
MAX_IMPORT_ROWS = 5000
MAX_XLSX_UNCOMPRESSED_BYTES = 25 * 1024 * 1024
MAX_XLSX_ARCHIVE_FILES = 1000


@dataclass(frozen=True)
class MovieImportError:
    """One actionable validation error tied to a workbook or data row."""

    row: int
    column: str | None
    message: str
    value: str | None = None

    def as_dict(self) -> dict:
        """Return only populated fields in API-ready form."""
        return {
            key: value
            for key, value in {
                'row': self.row,
                'column': self.column,
                'message': self.message,
                'value': self.value,
            }.items()
            if value is not None
        }


@dataclass(frozen=True)
class ValidatedMovieRow:
    """A parsed row whose TMDB reference and identifying fields agree."""

    row: int
    title: str
    release_year: int
    tmdb_id: int
    watched_date: date
    catalog_movie: DbMovie | None
    detail: dict | None


@dataclass
class MovieValidationResult:
    """All valid rows and errors found during the read-only validation pass."""

    rows: list[ValidatedMovieRow] = field(default_factory=list)
    errors: list[MovieImportError] = field(default_factory=list)


@dataclass(frozen=True)
class MovieImportOutcome:
    """The write-phase result for one validated source row."""

    row: int
    status: str
    tmdb_id: int
    movie_id: str
    title: str
    catalog_created: bool
    message: str

    def as_dict(self) -> dict:
        """Return the stable public shape for this row outcome."""
        return {
            'row': self.row,
            'status': self.status,
            'tmdb_id': self.tmdb_id,
            'movie_id': self.movie_id,
            'title': self.title,
            'catalog_created': self.catalog_created,
            'message': self.message,
        }


@dataclass
class MovieImportReport:
    """Per-row outcomes and their computed status counts."""

    outcomes: list[MovieImportOutcome] = field(default_factory=list)

    @property
    def summary(self) -> dict[str, int]:
        """Count each documented outcome status."""
        return {
            status: sum(outcome.status == status for outcome in self.outcomes)
            for status in ('imported', 'matched', 'skipped')
        }


def movie_template_csv() -> bytes:
    """Return an empty UTF-8 CSV template with the canonical header row."""
    output = io.StringIO(newline='')
    csv.writer(output).writerow(MOVIE_TEMPLATE_HEADERS)
    return output.getvalue().encode('utf-8')


def movie_template_xlsx() -> bytes:
    """Return an XLSX template with a data sheet and embedded instructions."""
    workbook = Workbook()
    movies = workbook.active
    movies.title = 'Movies'
    movies.append(MOVIE_TEMPLATE_HEADERS)
    movies.freeze_panes = 'A2'
    for cell in movies[1]:
        cell.font = Font(bold=True)
    for column, width in zip(('A', 'B', 'C', 'D'), (36, 16, 14, 20)):
        movies.column_dimensions[column].width = width

    instructions = workbook.create_sheet('Instructions')
    instructions.append(('column', 'required', 'accepted format', 'rule', 'example'))
    for cell in instructions[1]:
        cell.font = Font(bold=True)
    for column in MOVIE_TEMPLATE_COLUMNS:
        instructions.append(
            (
                column['name'],
                'yes',
                column['format'],
                column['description'],
                column['example'],
            )
        )
    instructions.append(
        (
            'file',
            'yes',
            f'at most {MAX_IMPORT_ROWS} movie rows',
            'Every row is validated before any history is written.',
            '',
        )
    )
    for column, width in zip(('A', 'B', 'C', 'D', 'E'), (20, 12, 38, 68, 24)):
        instructions.column_dimensions[column].width = width

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _file_error(message: str, value: str | None = None) -> MovieValidationResult:
    return MovieValidationResult(
        errors=[MovieImportError(row=1, column='file', message=message, value=value)]
    )


def _csv_rows(
    raw: bytes,
) -> tuple[list[str], list[tuple[int, dict]]] | MovieValidationResult:
    try:
        content = raw.decode('utf-8-sig')
    except UnicodeDecodeError:
        return _file_error('CSV must be UTF-8 text')

    try:
        reader = csv.reader(io.StringIO(content), strict=True)
        header = next(reader, None)
        if header is None:
            return _file_error('File is empty')
        headers = [str(value).strip() for value in header]
        rows = []
        for row_number, values in enumerate(reader, start=2):
            if row_number > MAX_IMPORT_ROWS + 1:
                return _file_error(
                    f'File exceeds the {MAX_IMPORT_ROWS}-row import limit'
                )
            if not any(str(value).strip() for value in values):
                continue
            rows.append((row_number, dict(zip(headers, values + [''] * len(headers)))))
    except csv.Error as exc:
        return _file_error(f'CSV could not be parsed: {exc}')
    return headers, rows


def _xlsx_rows(
    raw: bytes,
) -> tuple[list[str], list[tuple[int, dict]]] | MovieValidationResult:
    try:
        with ZipFile(io.BytesIO(raw)) as archive:
            files = archive.infolist()
            if len(files) > MAX_XLSX_ARCHIVE_FILES:
                return _file_error('XLSX contains too many internal files')
            if sum(item.file_size for item in files) > MAX_XLSX_UNCOMPRESSED_BYTES:
                return _file_error('XLSX expands beyond the safe processing limit')
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
        return _file_error(f'XLSX could not be parsed: {exc}')

    try:
        worksheet = workbook.active
        if worksheet.max_row > MAX_IMPORT_ROWS + 1:
            return _file_error(f'File exceeds the {MAX_IMPORT_ROWS}-row import limit')
        values = worksheet.iter_rows(values_only=True)
        header = next(values, None)
        if header is None:
            return _file_error('File is empty')
        headers = [str(value).strip() if value is not None else '' for value in header]
        rows = []
        for row_number, cells in enumerate(values, start=2):
            if not any(value is not None and str(value).strip() for value in cells):
                continue
            rows.append((row_number, dict(zip(headers, cells))))
        return headers, rows
    finally:
        workbook.close()


def _raw_rows(
    raw: bytes, filename: str | None
) -> tuple[list[str], list[tuple[int, dict]]] | MovieValidationResult:
    suffix = Path(filename or '').suffix.casefold()
    if suffix == '.csv':
        return _csv_rows(raw)
    if suffix == '.xlsx':
        return _xlsx_rows(raw)
    return _file_error('Upload a .csv or .xlsx movie template', suffix or None)


def _display_value(value) -> str | None:
    if value is None:
        return None
    return str(value).strip() or None


def _parse_year(value) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        year = value
    elif isinstance(value, float) and value.is_integer():
        year = int(value)
    else:
        raw = _display_value(value)
        if raw is None or not re.fullmatch(r'\d{4}', raw):
            return None
        year = int(raw)
    return year if 1000 <= year <= 9999 else None


def _parse_tmdb_id(value) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        tmdb_id = value
    elif isinstance(value, float) and value.is_integer():
        tmdb_id = int(value)
    else:
        raw = _display_value(value)
        if raw is None or not raw.isdigit():
            return None
        tmdb_id = int(raw)
    return tmdb_id if tmdb_id > 0 else None


def _parse_watched_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = _display_value(value)
    if raw is None:
        return None
    try:
        parsed = date.fromisoformat(raw)
    except ValueError:
        return None
    return parsed if parsed.isoformat() == raw else None


def _title_key(value: str | None) -> str:
    decomposed = unicodedata.normalize('NFKD', value or '')
    plain = ''.join(
        character for character in decomposed if not unicodedata.combining(character)
    )
    return ' '.join(re.findall(r'[a-z0-9]+', plain.casefold()))


def _canonical_year(movie: DbMovie | None, detail: dict | None) -> int | None:
    if movie is not None:
        if movie.year:
            return movie.year
        if movie.release_date:
            return movie.release_date.year
    if detail:
        year = detail.get('year')
        return int(year) if year else None
    return None


def validate_movie_import(  # pylint: disable=too-many-locals,too-many-branches,too-many-statements
    db: Session, raw: bytes, filename: str | None
) -> MovieValidationResult:
    """Parse and validate every row without mutating the database."""
    parsed = _raw_rows(raw, filename)
    if isinstance(parsed, MovieValidationResult):
        return parsed
    headers, raw_rows = parsed
    result = MovieValidationResult()

    duplicate_headers = sorted(
        {header for header in headers if header and headers.count(header) > 1}
    )
    for header in duplicate_headers:
        result.errors.append(
            MovieImportError(1, header, 'Column appears more than once')
        )
    for header in MOVIE_TEMPLATE_HEADERS:
        if header not in headers:
            result.errors.append(
                MovieImportError(1, header, 'Required column is missing')
            )
    if result.errors:
        return result
    if not raw_rows:
        return _file_error('File contains no movie rows')

    candidates = []
    seen_tmdb_ids = set()
    for row_number, row in raw_rows:
        row_error_count = len(result.errors)
        title = _display_value(row.get('title'))
        if not title:
            result.errors.append(
                MovieImportError(row_number, 'title', 'Title is required')
            )
        elif len(title) > 255:
            result.errors.append(
                MovieImportError(
                    row_number,
                    'title',
                    'Title must be 255 characters or fewer',
                    title,
                )
            )

        release_year = _parse_year(row.get('release_year'))
        if release_year is None:
            result.errors.append(
                MovieImportError(
                    row_number,
                    'release_year',
                    'Release year must be a four-digit year',
                    _display_value(row.get('release_year')),
                )
            )

        tmdb_id = _parse_tmdb_id(row.get('tmdb_id'))
        if tmdb_id is None:
            result.errors.append(
                MovieImportError(
                    row_number,
                    'tmdb_id',
                    'TMDB ID must be a positive integer',
                    _display_value(row.get('tmdb_id')),
                )
            )
        elif tmdb_id in seen_tmdb_ids:
            result.errors.append(
                MovieImportError(
                    row_number,
                    'tmdb_id',
                    'TMDB ID appears more than once in this file',
                    str(tmdb_id),
                )
            )
        else:
            seen_tmdb_ids.add(tmdb_id)

        watched_date = _parse_watched_date(row.get('watched_date'))
        if watched_date is None:
            result.errors.append(
                MovieImportError(
                    row_number,
                    'watched_date',
                    'Watched date must use YYYY-MM-DD',
                    _display_value(row.get('watched_date')),
                )
            )
        elif watched_date > date.today():
            result.errors.append(
                MovieImportError(
                    row_number,
                    'watched_date',
                    'Watched date cannot be in the future',
                    watched_date.isoformat(),
                )
            )

        if len(result.errors) == row_error_count:
            candidates.append((row_number, title, release_year, tmdb_id, watched_date))

    tmdb_ids = [candidate[3] for candidate in candidates]
    catalog_by_tmdb = {
        movie.tmdb: movie
        for movie in db.query(DbMovie).filter(DbMovie.tmdb.in_(tmdb_ids)).all()
    }
    for row_number, title, release_year, tmdb_id, watched_date in candidates:
        movie = catalog_by_tmdb.get(tmdb_id)
        detail = None
        canonical_title = movie.title if movie is not None else None
        canonical_year = _canonical_year(movie, None)
        if movie is None or not canonical_title or canonical_year is None:
            detail = get_movie_detail(tmdb_id)
            if detail:
                canonical_title = canonical_title or detail.get('title')
                canonical_year = canonical_year or _canonical_year(None, detail)

        row_error_count = len(result.errors)
        if movie is None and detail is None:
            result.errors.append(
                MovieImportError(
                    row_number,
                    'tmdb_id',
                    'TMDB ID does not resolve to a known movie',
                    str(tmdb_id),
                )
            )
        else:
            if canonical_title and _title_key(title) != _title_key(canonical_title):
                result.errors.append(
                    MovieImportError(
                        row_number,
                        'title',
                        f'Title does not match TMDB movie "{canonical_title}"',
                        title,
                    )
                )
            if canonical_year and release_year != canonical_year:
                result.errors.append(
                    MovieImportError(
                        row_number,
                        'release_year',
                        f'Release year does not match TMDB year {canonical_year}',
                        str(release_year),
                    )
                )
        if len(result.errors) == row_error_count:
            result.rows.append(
                ValidatedMovieRow(
                    row=row_number,
                    title=title,
                    release_year=release_year,
                    tmdb_id=tmdb_id,
                    watched_date=watched_date,
                    catalog_movie=movie,
                    detail=detail,
                )
            )
    return result


def _has_placed_movies(db: Session, user_pk: int) -> bool:
    return (
        db.query(DbUserMovie)
        .filter(
            DbUserMovie.user_id == user_pk,
            DbUserMovie.on_rankings.is_(True),
            DbUserMovie.rank.isnot(None),
        )
        .first()
        is not None
    )


def ingest_movie_import(
    db: Session, user_pk: int, rows: list[ValidatedMovieRow]
) -> MovieImportReport:
    """Ingest validated rows in one transaction and report every outcome."""
    report = MovieImportReport()
    has_placed_movies = _has_placed_movies(db, user_pk)
    try:
        for row in rows:
            movie = row.catalog_movie
            catalog_created = movie is None
            if movie is None:
                movie = DbMovie(
                    title=row.title,
                    tmdb=row.tmdb_id,
                    year=row.release_year,
                )
                if row.detail:
                    apply_detail_to_movie(movie, row.detail)
                db.add(movie)
                db.flush()

            tracker = (
                db.query(DbUserMovie)
                .filter(
                    DbUserMovie.user_id == user_pk,
                    DbUserMovie.movie_id == movie.pk,
                )
                .first()
            )
            if tracker is None:
                rank = 1 if not has_placed_movies else None
                tracker = DbUserMovie(
                    user_id=user_pk,
                    movie_id=movie.pk,
                    on_rankings=True,
                    on_watchlist=False,
                    rank=rank,
                    ranked_at=utc_now() if rank else None,
                    completed_at=row.watched_date,
                )
                db.add(tracker)
                if rank:
                    has_placed_movies = True
                outcome_status = 'imported'
                message = 'Added to movie history'
            else:
                changed = False
                if not tracker.on_rankings:
                    tracker.on_rankings = True
                    if not has_placed_movies:
                        tracker.rank = 1
                        tracker.ranked_at = utc_now()
                        has_placed_movies = True
                    else:
                        tracker.rank = None
                        tracker.ranked_at = None
                    changed = True
                if tracker.on_watchlist:
                    tracker.on_watchlist = False
                    changed = True
                if tracker.completed_at is None:
                    tracker.completed_at = row.watched_date
                    changed = True
                outcome_status = 'matched' if changed else 'skipped'
                message = (
                    'Updated existing movie history'
                    if changed
                    else 'Movie is already in history'
                )

            report.outcomes.append(
                MovieImportOutcome(
                    row=row.row,
                    status=outcome_status,
                    tmdb_id=row.tmdb_id,
                    movie_id=movie.id,
                    title=movie.title,
                    catalog_created=catalog_created,
                    message=message,
                )
            )
        db.commit()
    except Exception:
        db.rollback()
        raise
    return report
