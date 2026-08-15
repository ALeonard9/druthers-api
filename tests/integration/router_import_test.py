# pylint: disable=missing-module-docstring, missing-function-docstring
import io
from datetime import date, datetime, timezone
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.db.models_sandbox import DbMovie, DbUserMovie

HEADER = (
    'Title,Author,ISBN,ISBN13,My Rating,Number of Pages,Year Published,'
    'Original Publication Year,Exclusive Shelf,My Review,Date Read\n'
)

CSV = (
    HEADER
    + (
        'Dune,Frank Herbert,="0441172717",="9780441172719",5,412,1990,1965,'
        'read,A classic.,2020/01/02\n'
    )
    + 'Piranesi,Susanna Clarke,="",="9781635575637",0,245,2020,2020,to-read,,\n'
    + ',Nobody,="",="",0,,,,read,,\n'
)


def _auth(token: str) -> dict:
    return {'Authorization': f'Bearer {token}'}


def _upload(test_client: TestClient, token: str, content: str = CSV):
    return test_client.post(
        '/v1/users/me/import/goodreads',
        headers=_auth(token),
        files={'file': ('goodreads.csv', io.BytesIO(content.encode()), 'text/csv')},
    )


def _movie_upload(
    test_client: TestClient,
    token: str,
    content: bytes,
    filename: str = 'movies.csv',
):
    return test_client.post(
        '/v1/users/me/import/movies',
        headers=_auth(token),
        files={'file': (filename, io.BytesIO(content))},
    )


def _movie_csv(*rows: str) -> bytes:
    return (
        'title,release_year,tmdb_id,watched_date\n' + '\n'.join(rows) + '\n'
    ).encode()


def _tmdb_movie(title: str, tmdb_id: int, year: int) -> dict:
    return {
        'title': title,
        'tmdb': tmdb_id,
        'imdb': f'tt{tmdb_id:07d}',
        'year': year,
        'release_date': datetime(year, 1, 1),
        'runtime': 120,
        'poster_url': f'https://image.tmdb.org/t/p/w500/{tmdb_id}.jpg',
    }


@patch('app.services.goodreads_import.search_books', return_value=[])
@patch('app.services.goodreads_import.get_book_detail', return_value=None)
def test_goodreads_import_creates_books_and_trackers(
    mock_detail, _mock_search, test_client: TestClient
):
    token = test_client.first_user.token
    response = _upload(test_client, token)
    assert response.status_code == 200
    body = response.json()
    assert body['books_created'] == 2
    assert body['trackers_created'] == 2
    assert body['unplaced_read_book_ids'] == []
    assert body['skipped'] == [{'row': 4, 'reason': 'Missing title'}]

    books = test_client.get('/v1/users/me/books', headers=_auth(token)).json()
    by_title = {b['book']['title']: b for b in books}
    dune = by_title['Dune']
    assert dune['on_rankings'] is True
    assert dune['on_watchlist'] is False
    assert dune['rank'] == 1
    assert dune['completed_at'] == '2020-01-02'
    assert dune['book']['isbn'] == '9780441172719'
    assert dune['book']['year'] == 1965
    assert 'A classic.' in dune['notes']
    assert 'Goodreads rating: 5/5' in dune['notes']
    piranesi = by_title['Piranesi']
    assert piranesi['on_watchlist'] is True
    assert piranesi['on_rankings'] is False
    assert mock_detail.call_count == 2


@patch('app.services.goodreads_import.search_books', return_value=[])
@patch('app.services.goodreads_import.get_book_detail')
def test_goodreads_import_enriches_new_catalog_books(
    mock_detail, _mock_search, test_client: TestClient
):
    mock_detail.return_value = {
        'title': 'Dune',
        'isbn': '9780441172719',
        'authors': 'Frank Herbert',
        'year': 1965,
        'genre': 'Science fiction',
        'description': 'A desert planet and a great destiny.',
        'page_count': 604,
        'rating': 4.2,
        'language': 'eng',
        'poster_url': 'https://covers.openlibrary.org/b/id/123-L.jpg',
    }
    response = _upload(test_client, test_client.first_user.token)
    assert response.status_code == 200

    books = test_client.get(
        '/v1/users/me/books', headers=_auth(test_client.first_user.token)
    ).json()
    dune = next(book for book in books if book['book']['title'] == 'Dune')['book']
    assert dune['poster_url'] == 'https://covers.openlibrary.org/b/id/123-L.jpg'
    dune_id = dune['id']
    detail = test_client.get(
        f'/v1/books/{dune_id}', headers=_auth(test_client.first_user.token)
    ).json()
    assert detail['genre'] == 'Science fiction'
    assert detail['description'] == 'A desert planet and a great destiny.'
    assert detail['page_count'] == 604


@patch('app.services.goodreads_import.search_books', return_value=[])
@patch('app.services.goodreads_import.get_book_detail', return_value=None)
def test_goodreads_import_keeps_csv_fields_when_provider_misses(
    mock_detail, _mock_search, test_client: TestClient
):
    _upload(test_client, test_client.first_user.token)
    books = test_client.get(
        '/v1/users/me/books', headers=_auth(test_client.first_user.token)
    ).json()
    dune = next(book for book in books if book['book']['title'] == 'Dune')['book']
    assert dune['authors'] == 'Frank Herbert'
    assert dune['isbn'] == '9780441172719'
    assert dune['year'] == 1965
    assert dune['poster_url'] is None
    assert mock_detail.call_count == 2


@patch('app.services.goodreads_import.search_books', return_value=[])
@patch('app.services.goodreads_import.get_book_detail', return_value=None)
def test_goodreads_import_is_idempotent(
    mock_detail, _mock_search, test_client: TestClient
):
    token = test_client.first_user.token
    _upload(test_client, token)
    body = _upload(test_client, token).json()
    assert body['books_created'] == 0
    assert body['books_matched'] == 2
    assert body['trackers_created'] == 0
    assert body['trackers_updated'] == 0
    books = test_client.get('/v1/users/me/books', headers=_auth(token)).json()
    assert len(books) == 2
    dune = next(book for book in books if book['book']['title'] == 'Dune')
    assert dune['rank'] == 1
    assert dune['notes'] == 'A classic.\n\nGoodreads rating: 5/5'
    assert mock_detail.call_count == 2


@patch('app.services.goodreads_import.search_books', return_value=[])
@patch('app.services.goodreads_import.get_book_detail', return_value=None)
def test_goodreads_import_promotes_but_never_demotes(
    _mock_detail, _mock_search, test_client: TestClient
):
    token = test_client.first_user.token
    _upload(test_client, token)
    # Same file with Piranesi now read: watchlist → rankings
    promoted = CSV.replace('to-read', 'read')
    body = _upload(test_client, token, promoted).json()
    assert body['trackers_updated'] == 1
    books = test_client.get('/v1/users/me/books', headers=_auth(token)).json()
    piranesi = next(b for b in books if b['book']['title'] == 'Piranesi')
    assert piranesi['on_rankings'] is True
    assert piranesi['on_watchlist'] is False
    assert piranesi['rank'] is None


@patch('app.services.goodreads_import.search_books', return_value=[])
@patch('app.services.goodreads_import.get_book_detail', return_value=None)
def test_goodreads_import_leaves_all_read_rows_after_the_first_unplaced(
    _mock_detail, _mock_search, test_client: TestClient
):
    content = (
        HEADER
        + 'First,Author,="",="9780000000001",0,100,2020,2020,read,,\n'
        + 'Second,Author,="",="9780000000002",0,100,2020,2020,read,,\n'
        + 'Third,Author,="",="9780000000003",0,100,2020,2020,read,,\n'
    )
    body = _upload(test_client, test_client.first_user.token, content).json()
    books = test_client.get(
        '/v1/users/me/books', headers=_auth(test_client.first_user.token)
    ).json()
    by_title = {book['book']['title']: book for book in books}
    assert by_title['First']['rank'] == 1
    assert by_title['Second']['rank'] is None
    assert by_title['Third']['rank'] is None
    assert body['unplaced_read_book_ids'] == [
        by_title['Second']['book']['id'],
        by_title['Third']['book']['id'],
    ]


@patch('app.services.goodreads_import.search_books', return_value=[])
@patch('app.services.goodreads_import.get_book_detail', return_value=None)
def test_goodreads_import_into_populated_rankings_leaves_read_book_unplaced(
    _mock_detail, _mock_search, test_client: TestClient
):
    headers = _auth(test_client.first_user.token)
    book = test_client.post(
        '/v1/books', headers=headers, json={'title': 'Placed'}
    ).json()
    book_id = book['id']
    response = test_client.post(
        f'/v1/users/me/books/{book_id}', headers=headers, json={'on_rankings': True}
    )
    assert response.json()['rank'] == 1

    body = _upload(test_client, test_client.first_user.token).json()
    books = test_client.get('/v1/users/me/books', headers=headers).json()
    dune = next(book for book in books if book['book']['title'] == 'Dune')
    assert dune['rank'] is None
    assert body['unplaced_read_book_ids'] == [dune['book']['id']]


@patch('app.services.goodreads_import.search_books', return_value=[])
@patch('app.services.goodreads_import.get_book_detail', return_value=None)
@patch('app.services.goodreads_import.utc_now')
def test_goodreads_import_defaults_missing_read_date_to_import_day(
    mock_now, _mock_detail, _mock_search, test_client: TestClient
):
    mock_now.return_value = datetime(2026, 8, 14, 15, tzinfo=timezone.utc)
    content = HEADER + 'Undated,Author,="",="9780000000001",0,100,2020,2020,read,,\n'
    _upload(test_client, test_client.first_user.token, content)
    books = test_client.get(
        '/v1/users/me/books', headers=_auth(test_client.first_user.token)
    ).json()
    assert books[0]['completed_at'] == '2026-08-14'


@patch('app.services.goodreads_import.get_book_detail')
@patch('app.services.goodreads_import.search_books')
def test_goodreads_import_enriches_isbnless_rows_by_title_and_author(
    mock_search, mock_detail, test_client: TestClient
):
    mock_search.return_value = [
        {
            'title': 'The Great Gatsby',
            'authors': 'F. Scott Fitzgerald',
            'isbn': '9780743273565',
            'year': '1925',
            'poster_url': 'https://covers.openlibrary.org/b/id/7222246-L.jpg',
        },
        {
            'title': 'The Great Gatsby',
            'authors': 'Some Other Author',
            'isbn': '9780000000000',
            'year': '2000',
            'poster_url': 'https://covers.openlibrary.org/b/id/1-L.jpg',
        },
    ]
    mock_detail.side_effect = [
        None,
        {
            'title': 'The Great Gatsby',
            'authors': 'F. Scott Fitzgerald',
            'isbn': '9780743273565',
            'year': 1925,
            'page_count': 180,
            'poster_url': 'https://covers.openlibrary.org/b/id/7222246-L.jpg',
        },
    ]
    content = HEADER + 'The Great Gatsby,Wrong Author,="","",0,218,2004,2004,read,,\n'

    response = _upload(test_client, test_client.first_user.token, content)

    assert response.status_code == 200
    books = test_client.get(
        '/v1/users/me/books', headers=_auth(test_client.first_user.token)
    ).json()
    gatsby = books[0]['book']
    assert gatsby['title'] == 'The Great Gatsby'
    assert gatsby['authors'] == 'F. Scott Fitzgerald'
    assert gatsby['isbn'] == '9780743273565'
    assert gatsby['poster_url'] == 'https://covers.openlibrary.org/b/id/7222246-L.jpg'
    assert gatsby['page_count'] == 180
    mock_search.assert_called_once_with('The Great Gatsby Wrong Author')
    assert mock_detail.call_args_list[1].args == ('9780743273565',)


@patch('app.services.goodreads_import.get_book_detail')
@patch('app.services.goodreads_import.search_books')
def test_goodreads_import_retries_an_unresolved_isbn_by_title_and_author(
    mock_search, mock_detail, test_client: TestClient
):
    mock_search.return_value = [
        {
            'title': 'The Catcher in the Rye',
            'authors': 'J. D. Salinger',
            'isbn': '9780316769488',
            'year': '1951',
            'poster_url': 'https://covers.openlibrary.org/b/id/8231856-L.jpg',
        }
    ]
    mock_detail.side_effect = [
        None,
        {
            'title': 'The Catcher in the Rye',
            'authors': 'J. D. Salinger',
            'isbn': '9780316769488',
            'year': 1951,
            'poster_url': 'https://covers.openlibrary.org/b/id/8231856-L.jpg',
        },
    ]
    content = (
        HEADER
        + 'The Catcher in the Rye,J. D. Salinger,="","9780000000000",0,234,2014,1951,read,,\n'
    )

    _upload(test_client, test_client.first_user.token, content)

    books = test_client.get(
        '/v1/users/me/books', headers=_auth(test_client.first_user.token)
    ).json()
    catcher = books[0]['book']
    assert catcher['isbn'] == '9780316769488'
    assert catcher['poster_url'] == 'https://covers.openlibrary.org/b/id/8231856-L.jpg'
    mock_search.assert_called_once_with('The Catcher in the Rye J. D. Salinger')
    assert mock_detail.call_args_list[0].args == ('9780000000000',)
    assert mock_detail.call_args_list[1].args == ('9780316769488',)


@patch('app.services.goodreads_import.get_book_detail')
@patch('app.services.goodreads_import.search_books')
def test_goodreads_import_strips_series_qualifier_from_title_for_search(
    mock_search, mock_detail, test_client: TestClient
):
    mock_search.return_value = [
        {
            'title': "Harry Potter and the Philosopher's Stone",
            'authors': 'J.K. Rowling',
            'isbn': '9780747532743',
            'year': '1997',
            'poster_url': 'https://covers.openlibrary.org/b/id/10521270-L.jpg',
        }
    ]
    mock_detail.side_effect = [
        None,
        {
            'title': "Harry Potter and the Philosopher's Stone",
            'authors': 'J.K. Rowling',
            'isbn': '9780747532743',
            'year': 1997,
            'poster_url': 'https://covers.openlibrary.org/b/id/10521270-L.jpg',
        },
    ]
    content = (
        HEADER
        + '"Harry Potter and the Philosopher\'s Stone (Harry Potter, #1)",'
        + 'J.K. Rowling,="","9780000000000",0,223,1997,1997,read,,\n'
    )

    _upload(test_client, test_client.first_user.token, content)

    books = test_client.get(
        '/v1/users/me/books', headers=_auth(test_client.first_user.token)
    ).json()
    hp = books[0]['book']
    assert hp['title'] == "Harry Potter and the Philosopher's Stone"
    assert hp['poster_url'] == 'https://covers.openlibrary.org/b/id/10521270-L.jpg'
    mock_search.assert_called_once_with(
        "Harry Potter and the Philosopher's Stone J.K. Rowling"
    )


def test_goodreads_import_rejects_non_export(test_client: TestClient):
    body = _upload(
        test_client, test_client.first_user.token, 'just,some,columns\n1,2,3\n'
    ).json()
    assert body['books_created'] == 0
    assert body['skipped'][0]['reason'].startswith('Not a Goodreads export')


def test_goodreads_import_requires_auth(test_client: TestClient):
    response = test_client.post(
        '/v1/users/me/import/goodreads',
        files={'file': ('x.csv', io.BytesIO(b'Title\n'), 'text/csv')},
    )
    assert response.status_code == 401


def test_goodreads_import_rejects_oversized_file(test_client: TestClient):
    token = test_client.first_user.token
    content = b'x' * (5 * 1024 * 1024 + 1)
    response = test_client.post(
        '/v1/users/me/import/goodreads',
        headers=_auth(token),
        files={'file': ('huge.csv', io.BytesIO(content), 'text/csv')},
    )
    assert response.status_code == 413


def test_movie_templates_are_downloadable_and_documented(test_client: TestClient):
    headers = _auth(test_client.first_user.token)
    csv_response = test_client.get(
        '/v1/users/me/import/movies/template.csv', headers=headers
    )
    assert csv_response.status_code == 200
    assert csv_response.text == 'title,release_year,tmdb_id,watched_date\r\n'
    assert 'druthers-movies-template.csv' in csv_response.headers['content-disposition']

    xlsx_response = test_client.get(
        '/v1/users/me/import/movies/template.xlsx', headers=headers
    )
    assert xlsx_response.status_code == 200
    workbook = load_workbook(io.BytesIO(xlsx_response.content), read_only=True)
    assert workbook.sheetnames == ['Movies', 'Instructions']
    assert tuple(cell.value for cell in workbook['Movies'][1]) == (
        'title',
        'release_year',
        'tmdb_id',
        'watched_date',
    )
    workbook.close()

    schema = test_client.get('/openapi.json').json()
    operation = schema['paths']['/v1/users/me/import/movies']['post']
    assert (
        'Every field is required'
        in schema['paths']['/v1/users/me/import/movies/template.csv']['get'][
            'description'
        ]
    )
    assert 'no rows were written' in operation['responses']['422']['description']


@patch('app.services.generic_movie_import.get_movie_detail')
def test_movie_csv_import_creates_catalog_and_history_atomically(
    mock_detail, test_client: TestClient
):
    existing = DbMovie(title='The Matrix', tmdb=603, year=1999)
    test_client.test_db_session.add(existing)
    test_client.test_db_session.flush()
    mock_detail.return_value = _tmdb_movie('Spirited Away', 129, 2001)
    content = _movie_csv(
        'The Matrix,1999,603,2020-01-02',
        'Spirited Away,2001,129,2020-02-03',
    )

    response = _movie_upload(test_client, test_client.first_user.token, content)

    assert response.status_code == 200
    body = response.json()
    assert body['valid'] is True
    assert body['summary'] == {'imported': 2, 'matched': 0, 'skipped': 0}
    assert [row['status'] for row in body['rows']] == ['imported', 'imported']
    assert [row['catalog_created'] for row in body['rows']] == [False, True]
    trackers = (
        test_client.test_db_session.query(DbUserMovie)
        .filter(DbUserMovie.user_id == test_client.first_user.pk)
        .order_by(DbUserMovie.pk)
        .all()
    )
    assert len(trackers) == 2
    assert [tracker.completed_at for tracker in trackers] == [
        date(2020, 1, 2),
        date(2020, 2, 3),
    ]
    assert trackers[0].rank == 1
    assert trackers[1].rank is None
    spirited_away = (
        test_client.test_db_session.query(DbMovie).filter(DbMovie.tmdb == 129).one()
    )
    assert spirited_away.imdb == 'tt0000129'
    assert spirited_away.poster_url.endswith('/129.jpg')


@patch('app.services.generic_movie_import.get_movie_detail')
def test_movie_import_rejects_every_row_without_partial_writes(
    mock_detail, test_client: TestClient
):
    def detail(tmdb_id):
        if tmdb_id == 603:
            return _tmdb_movie('The Matrix', 603, 1999)
        return None

    mock_detail.side_effect = detail
    content = _movie_csv(
        'The Matrix,1999,603,2020-01-02',
        'Not a real movie,2020,999999999,2020-02-03',
    )

    response = _movie_upload(test_client, test_client.first_user.token, content)

    assert response.status_code == 422
    body = response.json()
    assert body['valid'] is False
    assert body['rows'] == []
    assert body['errors'] == [
        {
            'row': 3,
            'column': 'tmdb_id',
            'message': 'TMDB ID does not resolve to a known movie',
            'value': '999999999',
        }
    ]
    assert test_client.test_db_session.query(DbMovie).count() == 0
    assert test_client.test_db_session.query(DbUserMovie).count() == 0


def test_movie_import_matches_watchlist_then_skips_same_file(
    test_client: TestClient,
):
    movie = DbMovie(title='The Matrix', tmdb=603, year=1999)
    test_client.test_db_session.add(movie)
    test_client.test_db_session.flush()
    test_client.test_db_session.add(
        DbUserMovie(
            user_id=test_client.first_user.pk,
            movie_id=movie.pk,
            on_watchlist=True,
            on_rankings=False,
        )
    )
    test_client.test_db_session.flush()
    content = _movie_csv('The Matrix,1999,603,2020-01-02')

    first = _movie_upload(test_client, test_client.first_user.token, content)
    second = _movie_upload(test_client, test_client.first_user.token, content)

    assert first.status_code == 200
    assert first.json()['summary'] == {'imported': 0, 'matched': 1, 'skipped': 0}
    assert second.status_code == 200
    assert second.json()['summary'] == {'imported': 0, 'matched': 0, 'skipped': 1}
    trackers = test_client.test_db_session.query(DbUserMovie).all()
    assert len(trackers) == 1
    assert trackers[0].on_rankings is True
    assert trackers[0].on_watchlist is False
    assert trackers[0].completed_at == date(2020, 1, 2)
    assert trackers[0].rank == 1


@patch('app.services.generic_movie_import.get_movie_detail')
def test_movie_xlsx_import_accepts_native_excel_values(
    mock_detail, test_client: TestClient
):
    mock_detail.return_value = _tmdb_movie('Spirited Away', 129, 2001)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(('title', 'release_year', 'tmdb_id', 'watched_date'))
    worksheet.append(('Spirited Away', 2001, 129, date(2020, 2, 3)))
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    response = _movie_upload(
        test_client,
        test_client.first_user.token,
        output.getvalue(),
        'movies.xlsx',
    )

    assert response.status_code == 200
    assert response.json()['summary']['imported'] == 1
    tracker = test_client.test_db_session.query(DbUserMovie).one()
    assert tracker.completed_at == date(2020, 2, 3)


def test_movie_import_rejects_missing_columns_and_requires_auth(
    test_client: TestClient,
):
    invalid = _movie_upload(
        test_client,
        test_client.first_user.token,
        b'title,tmdb_id\nThe Matrix,603\n',
    )
    assert invalid.status_code == 422
    assert [error['column'] for error in invalid.json()['errors']] == [
        'release_year',
        'watched_date',
    ]

    unauthorized = test_client.post(
        '/v1/users/me/import/movies',
        files={'file': ('movies.csv', io.BytesIO(_movie_csv('X,2000,1,2020-01-01')))},
    )
    assert unauthorized.status_code == 401
