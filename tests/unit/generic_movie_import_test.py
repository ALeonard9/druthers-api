# pylint: disable=missing-module-docstring, missing-function-docstring
import io
from datetime import date
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from app.db.models_sandbox import DbMovie
from app.services.generic_movie_import import (
    MOVIE_TEMPLATE_HEADERS,
    movie_template_csv,
    movie_template_xlsx,
    validate_movie_import,
)


def test_download_templates_carry_the_real_contract():
    assert movie_template_csv().decode() == (
        'title,release_year,tmdb_id,watched_date\r\n'
    )

    workbook = load_workbook(io.BytesIO(movie_template_xlsx()), read_only=True)
    assert workbook.sheetnames == ['Movies', 'Instructions']
    assert tuple(cell.value for cell in workbook['Movies'][1]) == MOVIE_TEMPLATE_HEADERS
    instructions = list(workbook['Instructions'].values)
    assert instructions[1] == (
        'title',
        'yes',
        'text, maximum 255 characters',
        'Movie title. Must match the movie identified by tmdb_id.',
        'The Matrix',
    )
    assert instructions[-1][2] == 'at most 5000 movie rows'
    workbook.close()


@patch('app.services.generic_movie_import.get_movie_detail', return_value=None)
def test_validation_reports_each_bad_row_without_writes(mock_detail, test_db_session):
    movie = DbMovie(title='The Matrix', tmdb=603, year=1999)
    test_db_session.add(movie)
    test_db_session.flush()
    before_movies = test_db_session.query(DbMovie).count()
    content = (
        'title,release_year,tmdb_id,watched_date\n'
        'Wrong title,1999,603,2020-01-01\n'
        'Unknown,2000,999999999,2020-01-02\n'
        'Bad primitives,nineteen,abc,01/03/2020\n'
        'The Matrix,1999,603,2020-01-04\n'
    )

    result = validate_movie_import(test_db_session, content.encode(), 'movies.csv')

    assert result.rows == []
    assert {error.row for error in result.errors} == {2, 3, 4, 5}
    by_row = {}
    for error in result.errors:
        by_row.setdefault(error.row, []).append((error.column, error.message))
    assert by_row[2] == [('title', 'Title does not match TMDB movie "The Matrix"')]
    assert by_row[3] == [('tmdb_id', 'TMDB ID does not resolve to a known movie')]
    assert {column for column, _ in by_row[4]} == {
        'release_year',
        'tmdb_id',
        'watched_date',
    }
    assert by_row[5] == [('tmdb_id', 'TMDB ID appears more than once in this file')]
    assert test_db_session.query(DbMovie).count() == before_movies
    mock_detail.assert_called_once_with(999999999)


@patch('app.services.generic_movie_import.get_movie_detail')
def test_xlsx_dates_and_integral_ids_validate_against_tmdb(
    mock_detail, test_db_session
):
    mock_detail.return_value = {
        'title': 'Spirited Away',
        'tmdb': 129,
        'year': 2001,
    }
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(MOVIE_TEMPLATE_HEADERS)
    worksheet.append(('Spirited Away', 2001, 129.0, date(2020, 2, 3)))
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    result = validate_movie_import(test_db_session, output.getvalue(), 'movies.xlsx')

    assert result.errors == []
    assert len(result.rows) == 1
    row = result.rows[0]
    assert row.row == 2
    assert row.tmdb_id == 129
    assert row.watched_date == date(2020, 2, 3)
    assert row.detail['title'] == 'Spirited Away'


def test_validation_requires_every_documented_column(test_db_session):
    result = validate_movie_import(
        test_db_session,
        b'title,tmdb_id\nThe Matrix,603\n',
        'movies.csv',
    )

    assert [error.as_dict() for error in result.errors] == [
        {
            'row': 1,
            'column': 'release_year',
            'message': 'Required column is missing',
        },
        {
            'row': 1,
            'column': 'watched_date',
            'message': 'Required column is missing',
        },
    ]
